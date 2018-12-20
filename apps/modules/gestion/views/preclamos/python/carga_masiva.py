#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import os
import glob
import codecs
import string
import base64
import informixdb
import sqlite3
from datetime import datetime, date, time
from xlrd import open_workbook, cellname, xldate_as_tuple
import openpyxl


'''
Developed by @remicioluis
'''

def create_tmp():
    query_sqlite = 'CREATE TABLE ' + str(table_tmp) + ' ('
    query_sqlite+= ' id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL,'
    query_sqlite+= ' shipper TEXT,'
    query_sqlite+= ' producto TEXT,'
    query_sqlite+= ' ciclo TEXT,'
    query_sqlite+= ' cli_codigo TEXT,'
    query_sqlite+= ' mot_reclamo TEXT,'
    query_sqlite+= ' det_reclamo TEXT,'
    query_sqlite+= ' posicion INTEGER,'
    query_sqlite+= ' error_sql INTEGER,'
    query_sqlite+= ' error_info TEXT,'
    query_sqlite+= ' guia INTEGER'
    query_sqlite+= ' )'

    cursorsqlite.execute('drop table if exists ' + str(table_tmp))

    cursorsqlite.execute(query_sqlite)

def validar_reclamo(vp_linea, vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_cod_mot, vp_barra):
    query_string = "call scm_reclamo_valida_archivo("+str(vp_linea)+", "+str(vp_shipper)+", '"+str(vp_producto)+"', '"+str(vp_ciclo)+"', '"+str(vp_codigo)+"', '"+str(vp_cod_mot)+"', '"+str(vp_barra)+"');"
    cursor.execute(query_string)
    rs = cursor.fetchall()
    return rs

def insert_tmp(vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_mot_reclamo, vp_det_reclamo, row_index, error_sql, error_info, guia):
    query_sqlite = "insert into " + str(table_tmp) + " (shipper, producto, ciclo, cli_codigo, mot_reclamo, det_reclamo, posicion, error_sql, error_info, guia) "
    query_sqlite+= "values ("+str(vp_shipper)+", '"+str(vp_producto)+"', '"+str(vp_ciclo)+"', '"+str(vp_codigo)+"', '"+str(vp_mot_reclamo)+"', '"+str(vp_det_reclamo)+"', "+str(row_index + 1)+", "+str(error_sql)+", '"+str(error_info)+"', "+str(guia)+")"
    cursorsqlite.execute(query_sqlite)

param = base64.b64decode(sys.argv[1])
params = param.split('&')

table_tmp = 'tmp_' + str(params[2]) + '_' + str(params[1])

# fileName = 'Reclamos Masivo.xls'

fileName = str(params[0])

vp_linea = str(params[3])

# print filename

pathUpload = '/sistemas/weburbano/public_html/uploads/'
pathSqlite = '/sistemas/weburbano/db_sqlite/carga_masiva'

os.chdir(pathUpload)

'''
Parameters of connection - Informix
'''
os.environ['INFORMIXSERVER'] = 'ol_urbano'
conn = informixdb.connect('scm30@ol_urbano', user='informix', password='mh_1C$_2sX')
cursor = conn.cursor(rowformat = informixdb.ROW_AS_DICT)

'''
Parameters of connection  - Sqlite
'''
consqlite = sqlite3.connect(pathSqlite)
cursorsqlite = consqlite.cursor()

error = 0
msg_error = ''

if os.path.isfile(fileName):
    extension = os.path.splitext(fileName)[1][1:]
    if extension == 'xls' or extension == 'xlsx':

        create_tmp()

        if extension == 'xls':
            '''
            Other libraries for files .xls
            '''
            workbook = open_workbook(fileName)
            worksheet = workbook.sheet_by_index(0)
            
            for row_index in range(worksheet.nrows):
                if row_index != 0:
                    vp_shipper = int(worksheet.cell(row_index,0).value)
                    vp_producto = worksheet.cell(row_index,1).value
                    
                    try:
                        date_value = xldate_as_tuple(worksheet.cell(row_index,2).value,workbook.datemode)
                        vp_ciclo = date(*date_value[:3]).strftime('%d/%m/%Y')
                    except ValueError:
                        vp_ciclo = datetime.strptime(worksheet.cell(row_index,2).value, '%d/%m/%Y').strftime('%d/%m/%Y')

                    vp_codigo = worksheet.cell(row_index,3).value
                    vp_mot_reclamo = worksheet.cell(row_index,4).value
                    vp_det_reclamo = worksheet.cell(row_index,5).value
                    vp_barra = ''

                    rs = validar_reclamo(vp_linea, vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_mot_reclamo, vp_barra)

                    error_sql = int(rs[0]['error_sql'])
                    error_info = rs[0]['error_info']
                    guia = int(rs[0]['guia'])

                    if error_sql != 0:
                        error = 1

                    insert_tmp(vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_mot_reclamo, vp_det_reclamo, row_index, error_sql, error_info, guia)
            consqlite.commit()
        else:
            workbook = openpyxl.load_workbook(filename = fileName)
            worksheet = workbook.get_active_sheet()

            i = 0
            for cell in worksheet.rows:
                if i != 0:
                    vp_shipper = int(cell[0].value)
                    vp_producto = cell[1].value

                    # Validar en caso sea unicode
                    try:
                        vp_ciclo = cell[2].value.strftime('%d/%m/%Y')
                    except AttributeError:
                        vp_ciclo = datetime.strptime(cell[2].value, '%d/%m/%Y').strftime('%d/%m/%Y')

                    vp_codigo = cell[3].value
                    vp_mot_reclamo = cell[4].value
                    vp_det_reclamo = cell[5].value
                    vp_barra = ''

                    rs = validar_reclamo(vp_linea, vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_mot_reclamo, vp_barra)

                    error_sql = int(rs[0]['error_sql'])
                    error_info = rs[0]['error_info']
                    guia = int(rs[0]['guia'])

                    if error_sql != 0:
                        error = 1

                    insert_tmp(vp_shipper, vp_producto, vp_ciclo, vp_codigo, vp_mot_reclamo, vp_det_reclamo, i, error_sql, error_info, guia)

                i+=1
            consqlite.commit()
    else:
        msg_error = "Extension isn't valid."
else:
    msg_error = "File isn't exists."

'''
Closing connections
'''
cursor.close()
cursorsqlite.close()

'''
Out parameters
'''

print error
print msg_error
print params[2]